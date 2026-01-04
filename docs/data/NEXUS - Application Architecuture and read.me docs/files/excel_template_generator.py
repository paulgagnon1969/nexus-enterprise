"""
Excel Template Generator for Certified Payroll
Creates professionally formatted Excel templates with all 88 required columns
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_certified_payroll_template(output_path='/home/claude/certified_payroll_template.xlsx'):
    """Create a blank certified payroll template with all 88 columns"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Certified Payroll"
    
    # Define all 88 column headers in proper order
    headers = [
        # Identification (7)
        'payroll_number', 'project_code', 'contract_id', 'work_order', 
        'week_end_date', 'check_num', 'ssn', 'employee_ID', 'class_code',
        
        # Pay & Wages (7)
        'gross_employee_pay', 'all_projects', 'wages_paid_in_lieu_of_fringes', 
        'total_paid',
        
        # Daily Hours (8)
        'st_hrs_date1', 'st_hrs_date2', 'st_hrs_date3', 'st_hrs_date4', 
        'st_hrs_date5', 'st_hrs_date6', 'st_hrs_date7', 'Total_Hours_All_Projects',
        
        # Employee Benefits (8)
        'ep_haw', 'ep_pension', 'ep_vac_hol', 'ep_train', 'ep_all_other',
        'emp_ep_haw', 'emp_ep_pension', 'emp_ep_other',
        
        # Voluntary Contributions (4)
        'vol_cont_pension', 'vol_emp_pay_med', 
        'vol_cont_pension_rate', 'vol_cont_medical_rate',
        
        # Deductions (9)
        'dts_fed_tax', 'dts_fica', 'dts_medicare', 'dts_state_tax', 
        'dts_sdi', 'dts_dues', 'dts_savings', 'dts_other', 'dts_total',
        
        # Transportation (2)
        'trav_subs',
        
        # Rates (3)
        'pay_rate', 'OT_rate', '2OT_rate',
        
        # Rates & Fringes (3)
        'vac_hol_dues_rate', 'training_rate', 'in_lieu_payment_rate',
        
        # Notes (2)
        'prnotes', 'Payment_date',
        
        # Employee Info (11)
        'first_name', 'last_name', 'address1', 'address2', 
        'city', 'state', 'ZIP', 'phone', 
        'gender', 'ethnicity', 'Email',
        
        # Job Classification (5)
        'apprentice_id', 'craft_id', 'emp_status',
        
        # Compliance Flags (4)
        'vac_chk_box', 'fringe_paid_chk_box', 'date_hired', 'I9Verified',
        
        # Location (5)
        'work_county', 'Geographic_Ward', 'Geographic_Area', 
        'Congressional_District', 'State_Senate_District',
        
        # Demographics (3)
        'IsForeman', 'IsDisadvantaged', 'VeteranStatus',
        
        # Transportation & Other (7)
        'OtherDeductionNotes', 'num_exempt', 
        'DriversLicense', 'DriversLicenseState', 'Owner_Operator',
        
        # Other Deduction Details (3)
        'OD_Category', 'OD_Type', 'OD_Amount',
        
        # Compliance (2)
        'FringesProvidedByEmployer', 'LocalUnionNumber',
        
        # YTD (1)
        'YTD_SickPayTime'
    ]
    
    # Style definitions
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths based on content type
    column_widths = {
        'payroll_number': 12,
        'project_code': 12,
        'contract_id': 15,
        'work_order': 12,
        'week_end_date': 12,
        'check_num': 12,
        'ssn': 12,
        'employee_ID': 12,
        'class_code': 12,
        'first_name': 15,
        'last_name': 15,
        'address1': 30,
        'address2': 20,
        'city': 15,
        'state': 6,
        'ZIP': 10,
        'phone': 14,
        'email': 25,
        'prnotes': 40,
        'OtherDeductionNotes': 30
    }
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        
        # Set width
        if header in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[header]
        elif header.startswith('st_hrs_date') or header.startswith('ep_') or header.startswith('dts_'):
            ws.column_dimensions[col_letter].width = 10
        elif 'rate' in header.lower() or 'pay' in header.lower():
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 12
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Add sample data row with formulas
    sample_row = 2
    
    # Add instructions in a comment-like fashion
    ws['A3'] = "Instructions:"
    ws['A3'].font = Font(bold=True, italic=True)
    ws['A4'] = "• Enter payroll data starting at row 2"
    ws['A5'] = "• Hours columns: Enter actual hours worked each day (0-24)"
    ws['A6'] = "• Total_Hours_All_Projects: Sum of st_hrs_date1 through st_hrs_date7"
    ws['A7'] = "• All monetary fields should be in dollars and cents"
    ws['A8'] = "• SSN format: XXX-XX-XXXX"
    ws['A9'] = "• Date format: MM/DD/YY or YYYY-MM-DD"
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Created certified payroll template: {output_path}")
    print(f"  Total columns: {len(headers)}")
    return output_path


def create_populated_certified_payroll(db_path='certified_payroll.db',
                                       project_code='CBS',
                                       week_end_date='2025-12-19',
                                       output_path='/home/claude/certified_payroll_populated.xlsx'):
    """Create populated certified payroll Excel from database"""
    
    from payroll_db import CertifiedPayrollDB
    
    with CertifiedPayrollDB(db_path) as db:
        # Get payroll records
        records = db.get_weekly_certified_payroll(project_code, week_end_date)
        
        if not records:
            print(f"No payroll records found for {project_code} week ending {week_end_date}")
            return None
        
        # Format for Excel
        formatted_records = db.export_to_excel_format(records)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{project_code} {week_end_date}"
    
    # Get headers from first record
    headers = list(formatted_records[0].keys())
    
    # Style definitions
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Calibri', size=10)
    text_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_num, record in enumerate(formatted_records, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = record.get(header, '')
            
            # Handle different data types
            if value == '' or value is None:
                cell.value = ''
            elif isinstance(value, (int, float)):
                cell.value = value
                cell.alignment = number_alignment
                
                # Apply number formatting
                if 'hrs' in header or 'Hours' in header:
                    cell.number_format = '0.00'
                elif any(x in header for x in ['pay', 'rate', 'tax', 'fica', 
                                               'medicare', 'ep_', 'dts_', 'vol_',
                                               'wages', 'total', 'OD_Amount', 'trav_subs']):
                    cell.number_format = '$#,##0.00'
            else:
                cell.value = str(value)
                cell.alignment = text_alignment
            
            cell.font = data_font
            cell.border = thin_border
            
            # Alternate row colors
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Set column widths
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        
        if header in ['prnotes', 'OtherDeductionNotes', 'address1']:
            ws.column_dimensions[col_letter].width = 35
        elif header in ['first_name', 'last_name', 'Email']:
            ws.column_dimensions[col_letter].width = 18
        elif header in ['city', 'craft_id']:
            ws.column_dimensions[col_letter].width = 15
        elif header.startswith('st_hrs') or header.startswith('ep_') or header.startswith('dts_'):
            ws.column_dimensions[col_letter].width = 11
        else:
            ws.column_dimensions[col_letter].width = 13
    
    # Freeze panes
    ws.freeze_panes = 'A2'
    
    # Add a summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    # Summary headers
    summary_ws['A1'] = 'Certified Payroll Summary'
    summary_ws['A1'].font = Font(size=14, bold=True)
    
    summary_ws['A3'] = 'Project:'
    summary_ws['B3'] = project_code
    summary_ws['A4'] = 'Week Ending:'
    summary_ws['B4'] = week_end_date
    summary_ws['A5'] = 'Number of Employees:'
    summary_ws['B5'] = len(formatted_records)
    
    summary_ws['A7'] = 'Totals:'
    summary_ws['A7'].font = Font(bold=True)
    
    # Calculate totals
    total_hours = sum([r.get('Total_Hours_All_Projects', 0) or 0 for r in formatted_records])
    total_gross = sum([r.get('gross_employee_pay', 0) or 0 for r in formatted_records])
    total_paid = sum([r.get('total_paid', 0) or 0 for r in formatted_records])
    
    summary_ws['A8'] = 'Total Hours:'
    summary_ws['B8'] = total_hours
    summary_ws['B8'].number_format = '0.00'
    
    summary_ws['A9'] = 'Total Gross Pay:'
    summary_ws['B9'] = total_gross
    summary_ws['B9'].number_format = '$#,##0.00'
    
    summary_ws['A10'] = 'Total Paid (with benefits):'
    summary_ws['B10'] = total_paid
    summary_ws['B10'].number_format = '$#,##0.00'
    
    # Format summary sheet
    for row in range(3, 11):
        summary_ws[f'A{row}'].font = Font(bold=True)
    
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 20
    
    # Save
    wb.save(output_path)
    print(f"✓ Created populated certified payroll: {output_path}")
    print(f"  Project: {project_code}")
    print(f"  Week Ending: {week_end_date}")
    print(f"  Records: {len(formatted_records)}")
    print(f"  Total Hours: {total_hours:.2f}")
    print(f"  Total Cost: ${total_paid:.2f}")
    
    return output_path


if __name__ == "__main__":
    print("=" * 80)
    print("CERTIFIED PAYROLL EXCEL TEMPLATE GENERATOR")
    print("=" * 80)
    print()
    
    # Create blank template
    create_certified_payroll_template()
    print()
    
    # Create populated template from database
    create_populated_certified_payroll()
    
    print()
    print("=" * 80)
    print("TEMPLATES CREATED SUCCESSFULLY")
    print("=" * 80)
